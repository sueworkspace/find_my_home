#!/usr/bin/env python3
"""네이버 부동산 최저가 매물 찾기 에이전트 시스템

사용법:
    python main.py 래미안퍼스티지
    python main.py 래미안퍼스티지 반포자이
"""

import json
import sys
import os
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from crew import NaverRealEstateCrew


def parse_crew_output(crew_result) -> list[dict]:
    """CrewAI 출력에서 JSON 배열을 추출합니다."""
    raw = str(crew_result)

    # 1) 직접 JSON 파싱 시도
    try:
        parsed = json.loads(raw)
        if isinstance(parsed, list):
            return parsed
    except json.JSONDecodeError:
        pass

    # 2) 텍스트 내에서 JSON 배열 찾기
    start = raw.find("[")
    end = raw.rfind("]") + 1
    if start != -1 and end > start:
        try:
            parsed = json.loads(raw[start:end])
            if isinstance(parsed, list):
                return parsed
        except json.JSONDecodeError:
            pass

    # 3) ```json 코드블록 내에서 찾기
    import re
    code_block = re.search(r"```(?:json)?\s*(\[.*?\])\s*```", raw, re.DOTALL)
    if code_block:
        try:
            parsed = json.loads(code_block.group(1))
            if isinstance(parsed, list):
                return parsed
        except json.JSONDecodeError:
            pass

    print("WARNING: CrewAI 출력에서 JSON 데이터를 파싱할 수 없습니다.")
    print("Raw output (처음 500자):", raw[:500])
    return []


def write_excel(data: list[dict], apartment_names: str, output_dir: str = "output") -> str:
    """매물 데이터를 가격 오름차순으로 정렬하여 엑셀 파일로 저장합니다."""
    output_path = Path(output_dir)
    output_path.mkdir(exist_ok=True)

    # 가격 오름차순 정렬
    data_sorted = sorted(data, key=lambda x: int(x.get("price_manwon", 0) or 0))

    wb = Workbook()
    ws = wb.active
    ws.title = "최저가 매물 정리"

    # 헤더 스타일
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11, name="맑은 고딕")
    data_font = Font(size=10, name="맑은 고딕")
    price_font = Font(size=10, name="맑은 고딕", bold=True, color="CC0000")

    # 헤더 행
    headers = ["순위", "단지명", "주소", "평형", "층", "가격(만원)", "거래유형", "등록일"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 데이터 행
    for row_idx, item in enumerate(data_sorted, 2):
        rank = row_idx - 1
        ws.cell(row=row_idx, column=1, value=rank).font = data_font
        ws.cell(row=row_idx, column=2, value=item.get("complex_name", "")).font = data_font
        ws.cell(row=row_idx, column=3, value=item.get("address", "")).font = data_font
        ws.cell(row=row_idx, column=4, value=item.get("area_pyeong", "")).font = data_font
        ws.cell(row=row_idx, column=5, value=item.get("floor", "")).font = data_font

        price = int(item.get("price_manwon", 0) or 0)
        price_cell = ws.cell(row=row_idx, column=6, value=price)
        price_cell.font = price_font
        price_cell.number_format = "#,##0"

        ws.cell(row=row_idx, column=7, value=item.get("trade_type", "")).font = data_font
        ws.cell(row=row_idx, column=8, value=item.get("date", item.get("article_confirm_date", ""))).font = data_font

        # 가운데 정렬
        for col in [1, 4, 5, 6, 7, 8]:
            ws.cell(row=row_idx, column=col).alignment = Alignment(horizontal="center")

    # 요약 행
    summary_row = ws.max_row + 2
    ws.cell(row=summary_row, column=1, value="[ 요약 ]").font = Font(bold=True, size=11, name="맑은 고딕")
    ws.cell(row=summary_row, column=2, value=f"총 {len(data_sorted)}건").font = data_font
    if data_sorted:
        lowest = int(data_sorted[0].get("price_manwon", 0) or 0)
        highest = int(data_sorted[-1].get("price_manwon", 0) or 0)
        ws.cell(row=summary_row, column=3, value=f"최저가: {lowest:,}만원").font = Font(
            bold=True, color="CC0000", name="맑은 고딕"
        )
        ws.cell(row=summary_row, column=4, value=f"최고가: {highest:,}만원").font = data_font

    # 열 너비 자동 조정
    for col in range(1, len(headers) + 1):
        max_length = 0
        for row in range(1, ws.max_row + 1):
            val = ws.cell(row=row, column=col).value
            if val is not None:
                max_length = max(max_length, len(str(val)))
        ws.column_dimensions[get_column_letter(col)].width = min(max_length + 4, 45)

    # 최소 너비 보장
    ws.column_dimensions["A"].width = 6   # 순위
    ws.column_dimensions["B"].width = 20  # 단지명
    ws.column_dimensions["C"].width = 35  # 주소

    # 파일 저장
    safe_name = apartment_names.replace(" ", "_").replace("/", "_")
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"naver_부동산_최저가_{safe_name}_{timestamp}.xlsx"
    filepath = str(output_path / filename)
    wb.save(filepath)

    return filepath


def main():
    if len(sys.argv) < 2:
        print("사용법: python main.py <아파트이름> [아파트이름2 ...]")
        print("예시:   python main.py 래미안퍼스티지")
        print("        python main.py 래미안퍼스티지 반포자이")
        sys.exit(1)

    apartment_names = " ".join(sys.argv[1:])
    print(f"=== 네이버 부동산 최저가 매물 찾기 ===")
    print(f"검색 대상: {apartment_names}")
    print()

    # CrewAI 실행
    crew = NaverRealEstateCrew()
    result = crew.run(apartment_names)

    # 결과 파싱 및 엑셀 생성
    listings = parse_crew_output(result)

    if listings:
        filepath = write_excel(listings, apartment_names)
        print()
        print("=" * 50)
        print(f"엑셀 파일 저장 완료: {filepath}")
        print(f"총 매물 수: {len(listings)}건")
        prices = [int(l.get("price_manwon", 0) or 0) for l in listings if l.get("price_manwon")]
        if prices:
            print(f"최저가: {min(prices):,}만원")
            print(f"최고가: {max(prices):,}만원")
        print("=" * 50)
    else:
        print("매물 데이터를 찾지 못했거나 파싱에 실패했습니다.")
        print("CrewAI 원본 출력:")
        print(str(result)[:1000])


if __name__ == "__main__":
    main()
